from openpyxl import load_workbook
import os


# ============================================================
# NORMALIZAR CNPJ
# ============================================================

def normalizar_cnpj(cnpj):

    if not cnpj:
        return ""

    return (
        str(cnpj)
        .replace(".", "")
        .replace("/", "")
        .replace("-", "")
        .replace(" ", "")
        .strip()
    )


def formatar_cnpj(cnpj):

    cnpj = normalizar_cnpj(cnpj)

    if len(cnpj) != 14:
        return cnpj

    return (
        f"{cnpj[:2]}."
        f"{cnpj[2:5]}."
        f"{cnpj[5:8]}/"
        f"{cnpj[8:12]}-"
        f"{cnpj[12:]}"
    )


# ============================================================
# CONVERTER NÚMERO
# ============================================================

def converter_numero(valor):

    if valor is None or valor == "":
        return None

    valor = str(valor).strip()

    valor = valor.replace(".", "")
    valor = valor.replace(",", ".")

    try:
        return float(valor)

    except ValueError:
        return None


# ============================================================
# CARREGAR CLIENTES
# ============================================================

def carregar_clientes(ws_clientes):

    clientes = {}

    for linha in range(
        2,
        ws_clientes.max_row + 1
    ):

        cnpj = ws_clientes.cell(
            linha,
            1
        ).value

        razao = ws_clientes.cell(
            linha,
            2
        ).value

        endereco = ws_clientes.cell(
            linha,
            3
        ).value

        cnpj = normalizar_cnpj(cnpj)

        if not cnpj:
            continue

        clientes[cnpj] = {
            "razao": razao or "",
            "endereco": endereco or ""
        }

    return clientes


# ============================================================
# ABRIR PLANILHA
# ============================================================

def abrir_planilha(nome):

    if not os.path.exists(nome):

        raise FileNotFoundError(
            f"Não foi possível encontrar:\n{nome}"
        )

    wb = load_workbook(nome)

    # --------------------------------------------------------
    # Verifica REGISTROS
    # --------------------------------------------------------

    if "REGISTROS" not in wb.sheetnames:

        wb.close()

        raise ValueError(
            "A aba REGISTROS não existe."
        )

    # --------------------------------------------------------
    # Verifica CLIENTES
    # --------------------------------------------------------

    if "FATURADOS" not in wb.sheetnames:

        wb.close()

        raise ValueError(
            "A aba FATURADOS não existe."
        )

    # --------------------------------------------------------
    # Seleciona as abas
    # --------------------------------------------------------

    ws_registros = wb["REGISTROS"]

    ws_clientes = wb["FATURADOS"]

    # --------------------------------------------------------
    # Carrega os clientes em memória
    # --------------------------------------------------------

    clientes = carregar_clientes(
        ws_clientes
    )

    # --------------------------------------------------------
    # RETORNA OS 3 VALORES
    # --------------------------------------------------------

    return (
        wb,
        ws_registros,
        clientes
    )


# ============================================================
# ADICIONAR DANFE
# ============================================================

def adicionar_danfe(
    ws,
    danfe,
    clientes
):

    linha = ws.max_row + 1

    # --------------------------------------------------------
    # CNPJ
    # --------------------------------------------------------

    cnpj = normalizar_cnpj(
        danfe.cnpj
    )

    # --------------------------------------------------------
    # PROCURA CLIENTE
    # --------------------------------------------------------

    cliente = clientes.get(cnpj)

    if cliente:

        razao_social = cliente["razao"]

        endereco = cliente["endereco"]

        print(
            f"Cliente encontrado: {cnpj}"
        )

    else:

        razao_social = danfe.razao_social

        endereco = danfe.endereco

        print(
            f"CNPJ não cadastrado em FATURADOS: {cnpj}"
        )

    # --------------------------------------------------------
    # COLUNA A - CIDADE
    # --------------------------------------------------------

    ws.cell(
        linha,
        1
    ).value = danfe.cidade

    # --------------------------------------------------------
    # COLUNA B - DATA
    # --------------------------------------------------------

    ws.cell(
        linha,
        2
    ).value = danfe.data_emissao

    # --------------------------------------------------------
    # COLUNA C - RAZÃO SOCIAL
    # --------------------------------------------------------

    ws.cell(
        linha,
        3
    ).value = razao_social

    # --------------------------------------------------------
    # COLUNA D - ENDEREÇO
    # --------------------------------------------------------

    ws.cell(
        linha,
        4
    ).value = endereco

    # --------------------------------------------------------
    # COLUNA E - Nº DE NOTAS
    # --------------------------------------------------------

    ws.cell(
        linha,
        5
    ).value = 1

    # --------------------------------------------------------
    # COLUNA F - Nº DA NF
    # --------------------------------------------------------

    nf = ws.cell(
        linha,
        6
    )

    nf.value = converter_numero(
        danfe.numero_nf
    )

    nf.number_format = "0"

    # --------------------------------------------------------
    # COLUNA G - EMBARQUE
    # --------------------------------------------------------

    ws.cell(
        linha,
        7
    ).value = ""

    # --------------------------------------------------------
    # COLUNA H - CNPJ
    # --------------------------------------------------------

    ws.cell(
        linha,
        8
    ).value = formatar_cnpj(cnpj)

    # --------------------------------------------------------
    # COLUNA I - PESO
    # --------------------------------------------------------

    peso = ws.cell(
        linha,
        9
    )

    peso.value = converter_numero(
        danfe.peso_bruto
    )

    peso.number_format = "0.00"

    # --------------------------------------------------------
    # COLUNA J - VALOR
    # --------------------------------------------------------

    valor = ws.cell(
        linha,
        10
    )

    valor.value = converter_numero(
        danfe.valor_total
    )

    valor.number_format = "0.00"


# ============================================================
# SALVAR PLANILHA
# ============================================================

def salvar_planilha(
    wb,
    nome
):

    try:

        wb.save(nome)

        wb.close()

    except PermissionError:

        wb.close()

        raise PermissionError(
            "A planilha ROTA.xlsx está aberta. "
            "Feche a planilha e tente novamente."
        )